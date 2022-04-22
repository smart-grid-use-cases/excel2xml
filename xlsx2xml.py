#! /usr/bin/python3

import datetime
import re
import pyxb
import IEC62559
import sys
import xml.dom.minidom
from openpyxl import load_workbook


def cell(sheet_ranges, row, column):
    return sheet_ranges.cell(row=row, column=column).value


def cell_str(sheet_ranges, row, column):
    return str(sheet_ranges.cell(row=row, column=column).value)


def extract_area(sheet_ranges):
    area = IEC62559.Area()
    area.name = cell_str(sheet_ranges, 5, 3)
    return area


def extract_version(sheet_ranges):
    version = IEC62559.Version()
    author = IEC62559.Author()
    try:
        version.number = cell_str(sheet_ranges, 8, 3)
        tdate = str(sheet_ranges.cell(row=9, column=3).value).strip()
        if len(tdate) == 19:
            tdate = tdate[:-9]
            version.date = datetime.datetime.strptime(tdate, "%Y-%m-%d")
        else:
            version.date = datetime.datetime.strptime(tdate, "%d.%m.%Y")
        author.name = str(sheet_ranges.cell(row=10, column=3).value)
        version.changes = str(sheet_ranges.cell(row=11, column=3).value)
        version.approvalStatus = str(sheet_ranges.cell(row=12, column=3).value)
        version.Author.append(author)
    except Exception as e:
        print("Exception caught: " + str(e), file=sys.stderr)
        version = IEC62559.Version()
    return version


def extract_relobj(sheet_ranges):
    relobj = IEC62559.Objective()
    relobj.name = cell(sheet_ranges, 15, 3) or 'None'
    return relobj


def extract_bcase(sheet_ranges):
    bcase = IEC62559.Ref_BusinessCase()
    bcase.mRID = cell(sheet_ranges, 16, 3) or 'None'
    return bcase


def extract_narrative(sheet_ranges):
    narrative = IEC62559.Narrative()
    narrative.shortDescription = cell(sheet_ranges, 18, 3)
    narrative.completeDescription = cell(sheet_ranges, 19, 3)
    return narrative


def extract_kpis(sheet_ranges):
    kpi_list = []
    column = 3
    while True:
        kpi = IEC62559.KeyPerformanceIndicator()
        try:
            kpi.identifier = cell_str(sheet_ranges, 21, column)
            kpi.name = cell_str(sheet_ranges, 22, column)
            kpi.description = cell_str(sheet_ranges, 23, column)
            refobj = IEC62559.Ref_Objective()
            refobj.mRID = cell_str(sheet_ranges, 24, column)
            kpi.Objective.append(refobj)
            column = column + 1

            if kpi.name == 'None' and kpi.identifier == 'None':
                break
            else:
                kpi_list.append(kpi)

        except Exception as e:
            print("Exception caught: " + str(e), file=sys.stderr)

    return kpi_list


def extract_assumptions(sheet_ranges):
    assumptions = []
    column = 3
    while True:
        assumption = IEC62559.Assumption()
        try:
            assumption.description = cell_str(sheet_ranges, 26, column)
            if assumption.description == 'None':
                break
            else:
                assumptions.append(assumption)
            column = column + 1
        except Exception as e:
            print("Exception caught: " + str(e), file=sys.stderr)
    return assumptions


def extract_conditions(sheet_ranges):
    conditions = []
    column = 3
    while True:
        condition = IEC62559.Condition()
        try:
            condition.description = cell_str(sheet_ranges, 27, column)
            if condition.description == 'None':
                break
            else:
                conditions.append(condition)
            column = column + 1
        except Exception as e:
            print("Exception caught: " + str(e), file=sys.stderr)
    return conditions


def extract_ref_usecase(sheet_ranges):
    refusecase = IEC62559.Ref_UseCase()
    refusecase.mRID = cell(sheet_ranges, 29, 3) or 'None'
    return refusecase


def extract_remark(sheet_ranges):
    remark = IEC62559.Remark()
    try:
        remark.description = cell_str(sheet_ranges, 36, 3)
    except Exception as e:
        print("Exception caught: " + str(e), file=sys.stderr)
    return remark


def extract_drawings(sheet_ranges):

    def switch_drawingType(argument):
        switcher = {
            "other": IEC62559.DrawingClassification.other,
            "domain_overview": IEC62559.DrawingClassification.domain_overview,
            "use_case_overview":
            IEC62559.DrawingClassification.use_case_overview,
            "documentation": IEC62559.DrawingClassification.documentation,
            "scenarios_flowchart":
            IEC62559.DrawingClassification.scenarios_flowchart,
            "scenario_overview":
            IEC62559.DrawingClassification.scenario_overview,
            "activities_flowchart":
            IEC62559.DrawingClassification.activities_flowchart,
            "activity_overview":
            IEC62559.DrawingClassification.activity_overview,
            "business_objects_overview":
            IEC62559.DrawingClassification.business_objects_overview,
            "role_model": IEC62559.DrawingClassification.role_model,
            "None": IEC62559.DrawingClassification.other,
        }
        return switcher.get(argument, IEC62559.DrawingClassification.other)

    def switch_uriType(argument):
        switcher = {
            "UMLDiagram": IEC62559.ResourceType.UMLDiagram,
            "image": IEC62559.ResourceType.image,
            "None": IEC62559.ResourceType.image,
        }
        return switcher.get(argument, IEC62559.ResourceType.image)

    drawings = []
    index = 0

    while index >= 0:
        try:
            name = cell_str(sheet_ranges, 38, 3 + index)
            drawingType = cell_str(sheet_ranges, 39, 3 + index)
            uriType = cell_str(sheet_ranges, 40, 3 + index)
            uri = cell_str(sheet_ranges, 41,
                           3 + index).strip().replace(' ', '_')

            if uri.endswith('.pdf'):
                uri = uri + '.png'

            if name != 'None' and uri != 'None':
                drawing = IEC62559.Drawing()
                resourcestr = IEC62559.Resource_String(uri)
                drawing.name = name  #cell_str(sheet_ranges, 38, 3)

                drawing.drawingType = switch_drawingType(
                    drawingType)  #cell_str(sheet_ranges, 39, 3)
                resourcestr.type = switch_uriType(
                    uriType
                )  #IEC62559.ResourceType.UMLDiagram #cell_str(sheet_ranges, 40, 3)
                drawing.URI = resourcestr
                drawings.append(drawing)
                index = index + 1
            else:
                #return None
                index = -1

        except Exception as e:
            print("Exception caught: " + str(e), file=sys.stderr)
            drawing = None
            index = -1

    return drawings


def extract_actorgroupings(sheet_ranges):
    actorgroupings = []
    column = 3
    while True and column < 20:
        try:
            actorgrouping = IEC62559.ActorGrouping()
            actorgrouping.name = cell_str(sheet_ranges, 44, column)
            actorgrouping.description = cell_str(sheet_ranges, 45, column)
            column = column + 1
            if actorgrouping.name == 'None' and actorgrouping.description == 'None':
                break
            else:
                actorgroupings.append(actorgrouping)
        except Exception as e:
            print("Exception caught: " + str(e), file=sys.stderr)
    return actorgroupings


def extract_actors(sheet_ranges):
    #actors = []
    actors = {}
    column = 3
    while True and column < 20:
        try:
            actor = IEC62559.Actor()
            actor.mRID = cell_str(sheet_ranges, 46, column)
            actor.name = cell_str(sheet_ranges, 46, column)
            actor.type = cell_str(sheet_ranges, 47, column)
            actor.description = cell_str(sheet_ranges, 48, column)
            column = column + 1
            if actor.name == 'None' and actor.description == 'None':
                break
            else:
                actors[actor.name]=actor
        except Exception as e:
            print("Exception caught: " + str(e), file=sys.stderr)
    return actors


def extract_references(sheet_ranges):
    references = []
    column = 3
    while True and column < 20:
        try:
            reference = IEC62559.Reference()
            reference.name = cell_str(sheet_ranges, 51, column)
            reference.number = cell_str(sheet_ranges, 52, column)
            reference.type = cell_str(sheet_ranges, 53, column)
            reference.description = cell_str(sheet_ranges, 54, column)
            reference.status = cell_str(sheet_ranges, 55, column)
            reference.impact = cell_str(sheet_ranges, 56, column)
            reference.originatorOrganization = cell_str(
                sheet_ranges, 57, column)
            reference.link = cell_str(sheet_ranges, 58, column)
            column = column + 1
            if reference.name == 'None' and reference.type == 'None':
                break
            else:
                references.append(reference)
        except Exception as e:
            print("Exception caught: " + str(e), file=sys.stderr)
    return references


def extract_scenarios(sheet_ranges):
    scenarios = []
    column = 3
    while True:
        scenario = IEC62559.Scenario()
        try:
            scenario.number = cell_str(sheet_ranges, 62, column)
            scenario.name = cell_str(sheet_ranges, 63, column)
            scenario.description = cell_str(sheet_ranges, 64, column)
            triggevent = IEC62559.TriggeringEvent()
            triggevent.description = cell_str(sheet_ranges, 66, column)
            precondition = IEC62559.Condition()
            precondition.description = cell_str(sheet_ranges, 67, column)
            postcondition = IEC62559.Condition()
            postcondition.description = cell_str(sheet_ranges, 68, column)
            scenario.TriggeringEvent.append(triggevent)
            scenario.Precondition.append(precondition)
            scenario.Postcondition.append(postcondition)

            if scenario.name == 'None' and scenario.description == 'None' and scenario.number == 'None':
                break
            else:
                scenarios.append(scenario)
            column = column + 1
        except Exception as e:
            print("Exception caught: " + str(e), file=sys.stderr)
    return scenarios


def extract_activities(sheet_ranges,actors,requirements):
    activities = []
    column = 3
    while True:
        activity = IEC62559.Activity()
        try:
            activity.number = cell_str(sheet_ranges, 71, column)
            activity.event = cell_str(sheet_ranges, 72, column)
            activity.name = cell_str(sheet_ranges, 73, column)
            activity.description = cell_str(sheet_ranges, 74, column)
            activity.service = cell_str(sheet_ranges, 75, column)
            activity.step_no = cell_str(sheet_ranges, 79, column)
            
            for e in cell_str(sheet_ranges, 76, column).split(","):
              if e.strip() in actors:
                ra = IEC62559.Ref_Actor()
                ra.mRID=actors[e.strip()].mRID
                activity.InformationProducer.append(ra)

            for e in cell_str(sheet_ranges, 77, column).split(","):
              if e.strip() in actors:
                ra = IEC62559.Ref_Actor()
                ra.mRID=actors[e.strip()].mRID
                activity.InformationReceiver.append(ra)
            
            for e in cell_str(sheet_ranges, 78, column).replace(","," ").replace("\"","").split():
              if e.strip() in requirements:
                rr = IEC62559.Ref_Requirement()
                rr.mRID=requirements[e.strip()].mRID
                activity.Requirement.append(rr)
            #activity.mRID = requirements[cell_str(sheet_ranges, 78, column)]
            column = column + 1

            if activity.description == 'None' and activity.name == 'None' and activity.number == 'None':
                break
            else:
                activities.append(activity)
        except Exception as e:
            print("Exception caught: " + str(e), file=sys.stderr)
    return activities


def match_activities_to_scenarios(activities, scenarios):
    for activity in activities:
        for scenario in scenarios:
            if scenario.number == activity.step_no:
                scenario.Activity.append(activity)


def extract_requirements(sheet_ranges):
    #requirements = []
    requirements = {}
    column = 3
    while True:
        requirement = IEC62559.Requirement()
        try:
            requirement.mRID = cell_str(sheet_ranges, 81, column) if cell_str(sheet_ranges, 81, column) != 'None' else cell_str(sheet_ranges, 82, column)
            requirement.identifier = cell_str(sheet_ranges, 82, column)
            requirement.name = cell_str(sheet_ranges, 83, column)
            requirement.description = cell_str(sheet_ranges, 84, column)
            column = column + 1

            if requirement.description == 'None' and requirement.identifier == 'None':
                break
            else:
                #requirements.append(requirement)
                requirements[requirement.identifier]=requirement
        except Exception as e:
            print("Exception caught: " + str(e), file=sys.stderr)
    return requirements


def extract_usecase(sheet_ranges, actors, requirements):
    column = 3
    usecase = IEC62559.UseCase()
    usecase.identifier = str(cell(sheet_ranges, 4, column))
    usecase.name = cell(sheet_ranges, 6, column)
    usecase.scope = cell(sheet_ranges, 14, column)
    usecase.RelatedObjective.append(extract_relobj(sheet_ranges))
    usecase.BusinessCase.append(extract_bcase(sheet_ranges))
    usecase.Narrative = extract_narrative(sheet_ranges)
    usecase.RelatedUseCase.append(extract_ref_usecase(sheet_ranges))
    usecase.levelOfDepth = cell(sheet_ranges, 30, column)
    usecase.prioritization = cell(sheet_ranges, 31, column)
    usecase.classification = cell(sheet_ranges, 32, column)
    usecase.nature = cell(sheet_ranges, 33, column)
    usecase.keywords = cell(sheet_ranges, 34, column)
    usecase.Version.append(extract_version(sheet_ranges))
    for drawing in extract_drawings(sheet_ranges):
        usecase.Drawing.append(drawing)
    for kpi in extract_kpis(sheet_ranges):
        usecase.KeyPerformanceIndicator.append(kpi)
    for assumption in extract_assumptions(sheet_ranges):
        usecase.Assumption.append(assumption)
    for condition in extract_conditions(sheet_ranges):
        usecase.Prerequisite.append(condition)
    usecase.Remark.append(extract_remark(sheet_ranges))
    for actorgrouping in extract_actorgroupings(sheet_ranges):
        usecase.append(actorgrouping)
    for reference in extract_references(sheet_ranges):
        usecase.Reference.append(reference)
    scenarios = extract_scenarios(sheet_ranges)
    for scenario in scenarios:
        usecase.Scenario.append(scenario)
    activities = extract_activities(sheet_ranges, actors, requirements)
    match_activities_to_scenarios(activities, scenarios)
    return usecase


def main():

    if len(sys.argv) > 1:
        filename = str(sys.argv[1])
    else:
        print("No arguments introduced", file=sys.stderr)

    try:
        wb = load_workbook(filename)
    except:
        print("File does not exist!", file=sys.stderr)
        pass

    sheet_list = wb.sheetnames
    usecaserep = IEC62559.UseCaseRepository()
    usecaselib = IEC62559.UseCaseLibrary()
    arealib = IEC62559.AreaLibrary()
    actorlib = IEC62559.ActorLibrary()
    reqlib = IEC62559.RequirementLibrary()
    reqcat = IEC62559.RequirementCategory()
  
    actors = {}
    requirements = {}
    usecase = None
    for sheet in wb:
        sheet_ranges = wb[sheet.title]
        actors = extract_actors(sheet_ranges)
        requirements = extract_requirements(sheet_ranges)

        usecase = extract_usecase(sheet_ranges,actors,requirements)
        arealib.Area.append(extract_area(sheet_ranges))
        usecaserep.AreaLibrary = arealib
        
        for actor in list(actors.values()):
            actorlib.append(actor)
        for requirement in list(requirements.values()):
            reqcat.Requirement.append(requirement)
        break

    usecaselib.UseCase.append(usecase)
    usecaselib.name = "UCL_name"
    usecaserep.UseCaseLibrary = usecaselib
    usecaserep.name = "UCR_name"
    usecaserep.ActorLibrary.append(actorlib)

    reqcat.name = "Req_Name"  #sustituir por cell
    reqcat.identifier = "Req_ID"  #sustituir por cell
    reqlib.append(reqcat)
    usecaserep.RequirementLibrary = reqlib

    print(
        xml.dom.minidom.parseString(
            usecaserep.toxml("utf-8", element_name='UseCaseRepository').decode(
                'utf-8')).toprettyxml())


#Python3
main()
